import os
import re
import sys

# Configure stdout/stderr to use UTF-8
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

report_path = "inspection_report.txt"
report = []

def log(msg):
    report.append(msg)
    try:
        print(msg)
    except Exception:
        # Fallback to ascii representation if printing still fails
        print(msg.encode('ascii', errors='replace').decode('ascii'))

def inspect_excel(file_path):
    log(f"\n============================================================\nInspecting Excel: {os.path.basename(file_path)}\n============================================================")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        log(f"Sheet names: {wb.sheetnames}")
        for name in wb.sheetnames:
            sheet = wb[name]
            log(f"\nSheet: {name} (Rows: {sheet.max_row}, Cols: {sheet.max_column})")
            for r in range(1, sheet.max_row + 1):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(15, sheet.max_column + 1))]
                if any(v is not None for v in row_vals):
                    cleaned_vals = [str(v) if v is not None else "" for v in row_vals]
                    while cleaned_vals and cleaned_vals[-1] == "":
                        cleaned_vals.pop()
                    if cleaned_vals:
                        log(f"Row {r:02d}: {cleaned_vals}")
    except Exception as e:
        log(f"Error inspecting Excel: {e}")

def inspect_doc(file_path):
    log(f"\n============================================================\nInspecting Word Doc: {os.path.basename(file_path)}\n============================================================")
    try:
        with open(file_path, 'rb') as f:
            content = f.read()
        
        # Word .doc files typically store text in UTF-16LE or 8-bit ANSI formats.
        # Let's extract UTF-16LE text first
        utf16_runs = []
        # Find sequences of printable/common Vietnamese characters in UTF-16LE
        # In UTF-16LE, a character is 2 bytes. Printable ASCII is byte, 0x00.
        # Let's extract any sequence of (printable ASCII or Vietnamese unicode) followed by 0x00
        # or other common Vietnamese characters.
        # Let's use a simpler heuristic: look for any block of even bytes where second byte is mostly 0x00 or 0x01/0x03/0x04/0x1e/0x1f (which cover Vietnamese unicode ranges)
        # Or let's just decode the entire file in UTF-16LE with errors='ignore' and extract readable sequences.
        decoded_16 = content.decode('utf-16le', errors='ignore')
        # Clean up and find words
        cleaned_16 = re.findall(r'[\w\s\-\.\,\:\/\(\)\%\+\=\[\]\u00C0-\u1EF9]{4,}', decoded_16)
        log("--- UTF-16LE Extracted Paragraphs/Segments ---")
        logged_count = 0
        for chunk in cleaned_16:
            chunk = re.sub(r'\s+', ' ', chunk).strip()
            if len(chunk) > 10 and not chunk.isspace():
                # Filter out obvious binary noise
                if any(c.isalnum() for c in chunk):
                    log(f"Seg16 {logged_count:03d}: {chunk}")
                    logged_count += 1
                    if logged_count >= 150:
                        break

        # Also try UTF-8/ASCII decode
        decoded_8 = content.decode('utf-8', errors='ignore')
        cleaned_8 = re.findall(r'[\w\s\-\.\,\:\/\(\)\%\+\=\[\]\u00C0-\u1EF9]{6,}', decoded_8)
        log("\n--- UTF-8 Extracted Paragraphs/Segments ---")
        logged_count = 0
        for chunk in cleaned_8:
            chunk = re.sub(r'\s+', ' ', chunk).strip()
            if len(chunk) > 12 and not chunk.isspace():
                # Avoid logging segments that look like random noise or duplicate UTF-16 runs
                if any(c.isalnum() for c in chunk) and not any(c in chunk for c in ['Adobe', 'Photoshop', 'ICC_PROFILE']):
                    log(f"Seg8 {logged_count:03d}: {chunk}")
                    logged_count += 1
                    if logged_count >= 100:
                        break

    except Exception as e:
        log(f"Error inspecting Doc: {e}")

dir_path = r"C:\Users\DAT\Downloads\MẪU"
inspect_excel(os.path.join(dir_path, "Bien ban ban giao.xlsx"))
inspect_excel(os.path.join(dir_path, "Bien ban nghiem thu.xlsx"))
inspect_doc(os.path.join(dir_path, "PR.doc"))

# Write report to UTF-8 file
with open(report_path, "w", encoding="utf-8") as f:
    f.write("\n".join(report))
print(f"\nReport written to {report_path}")

